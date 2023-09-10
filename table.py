import subprocess
from openpyxl import Workbook
from openpyxl.styles import Border, Side
import random


#ТАБЛИЦЫ
icebreakers = ["50 лет Победы", "Ямал", "Таймыр", "Вайгач"]
requests = [
   {"name": "Штурман Альбанов", "imo": 9752084, "ice_class": "Arc7", "speed": 15,
    "start_point": "точка в Баренцевом море", "end_point": "Саббета 3",
    "start_time": "01.01.21 23:30", "end_time": "05.01.21 5:00", "Day": 1},


   {"name": "Штурман Альбанов", "imo": 9752084, "ice_class": "Arc7", "speed": 15,
    "start_point": "Саббета 3", "end_point": "точка в Баренцевом море",
    "start_time": "06.01.21 11:00", "end_time": "10.01.21 18:00","Day": 6},


   {"name": "Штурман Альбанов", "imo": 9752084, "ice_class": "Arc7", "speed": 15,
    "start_point": "точка в Баренцевом море", "end_point": "Саббета 3",
    "start_time": "12.01.21 18:00", "end_time": "16.01.21 22:00","Day": 12},


   {"name": "Штурман Альбанов", "imo": 9752084, "ice_class": "Arc7", "speed": 15,
    "start_point": "Саббета 3", "end_point": "точка в Баренцевом море",
    "start_time": "17.01.21 6:00", "end_time": "21.01.21 20:45","Day": 17},


   {"name": "Штурман Альбанов", "imo": 9752084, "ice_class": "Arc7", "speed": 15,
    "start_point": "точка в Баренцевом море", "end_point": "Саббета 3",
    "start_time": "21.01.21 9:00", "end_time": "26.01.21 22:00","Day": 21},


   {"name": "Штурман Альбанов", "imo": 9752084, "ice_class": "Arc7", "speed": 15,
    "start_point": "Саббета 3", "end_point": "точка в Баренцевом море",
    "start_time": "28.01.21 1:00", "end_time": "02.02.21 19:00", "Day": 28},


   {"name": "Штурман Кошелев", "imo": 9759939, "ice_class": "Arc7", "speed": 15,
    "start_point": "Саббета 3", "end_point": "точка в Баренцевом море",
    "start_time": "02.01.21 20:30", "end_time": "06.01.21 7:00","Day": 2},


   {"name": "Штурман Кошелев", "imo": 9759939, "ice_class": "Arc7", "speed": 15,
    "start_point": "точка в Баренцевом море", "end_point": "Саббета 3",
    "start_time": "07.01.21 19:00", "end_time": "11.01.21 3:20","Day": 7},


   {"name": "Штурман Кошелев", "imo": 9759939, "ice_class": "Arc7", "speed": 15,
    "start_point": "Саббета 3", "end_point": "точка в Баренцевом море",
    "start_time": "13.01.21 16:00", "end_time": "17.01.21 14:00","Day": 13},


   {"name": "Штурман Кошелев", "imo": 9759939, "ice_class": "Arc7", "speed": 15,
    "start_point": "точка в Баренцевом море", "end_point": "Саббета 3",
    "start_time": "17.01.21 13:00", "end_time": "21.01.21 16:20","Day": 17},


   {"name": "Штурман Кошелев", "imo": 9759939, "ice_class": "Arc7", "speed": 15,
    "start_point": "Саббета 3", "end_point": "точка в Баренцевом море",
    "start_time": "22.01.21 7:30", "end_time": "26.01.21 19:00","Day": 22},


   {"name": "Штурман Кошелев", "imo": 9759939, "ice_class": "Arc7", "speed": 15,
    "start_point": "точка в Баренцевом море", "end_point": "Саббета 3",
    "start_time": "28.01.21 2:00", "end_time": "03.02.21 19:00","Day": 28},


   {"name": "Лагорта", "imo": 9194012, "ice_class": "Arc5", "speed": 15,
    "start_point": "точка в Баренцевом море", "end_point": "Саббета 3",
    "start_time": "27.12.20 23:00", "end_time": "07.01.21 17:00","Day": 27},


   {"name": "Лагорта", "imo": 9194012, "ice_class": "Arc5", "speed": 15,
    "start_point": "Саббета 3", "end_point": "точка в Баренцевом море",
    "start_time": "20.01.21 15:00", "end_time": "30.01.21 22:00","Day": 20},


   {"name": "Михаил Лазарев", "imo": 9837547, "ice_class": "Arc7", "speed": 15,
    "start_point": "Саббета 2", "end_point": "точка в Баренцевом море",
    "start_time": "14.01.21 22:00", "end_time": "18.01.21 10:00","Day": 14},


   {"name": "Михаил Лазарев", "imo": 9837547, "ice_class": "Arc7", "speed": 15,
    "start_point": "точка в Баренцевом море", "end_point": "Саббета 2",
    "start_time": "19.01.21 12:00", "end_time": "23.01.21 12:00"},


   {"name": "Юрий Кучиев", "imo": 9804033, "ice_class": "Arc7", "speed": 14,
    "start_point": "Саббета 1", "end_point": "точка в Баренцевом море",
    "start_time": "07.01.21 0:00", "end_time": "10.01.21 14:00","Day": 7},


   {"name": "Юрий Кучиев", "imo": 9804033, "ice_class": "Arc7", "speed": 14,
    "start_point": "точка в Баренцевом море", "end_point": "Саббета 1",
    "start_time": "26.01.21 22:00", "end_time": "31.01.21 20:00","Day": 26},


   {"name": "Эдуард Толль", "imo": 9750696, "ice_class": "Arc7", "speed": 19,
    "start_point": "точка в Баренцевом море", "end_point": "Саббета 1",
    "start_time": "01.01.21 17:30", "end_time": "04.01.21 5:00","Day": 1},


   {"name": "Эдуард Толль", "imo": 9750696, "ice_class": "Arc7", "speed": 19,
    "start_point": "Саббета 1", "end_point": "точка в Баренцевом море",
    "start_time": "07.01.21 17:00", "end_time": "11.01.21 12:30","Day": 7},


   {"name": "Борис Давыдов", "imo": 9768394, "ice_class": "Arc7", "speed": 19,
    "start_point": "точка в Баренцевом море", "end_point": "Саббета 1",
    "start_time": "08.01.21 22:00", "end_time": "10.01.21 14:15","Day": 8},


   {"name": "Владимир Воронин", "imo": 9750737, "ice_class": "Arc7", "speed": 19,
    "start_point": "точка в Баренцевом море", "end_point": "Саббета 1",
    "start_time": "16.01.21 18:00", "end_time": "18.01.21 19:00","Day": 16},


   {"name": "Владимир Воронин", "imo": 9750737, "ice_class": "Arc7", "speed": 19,
    "start_point": "Саббета 1", "end_point": "точка в Баренцевом море",
    "start_time": "26.01.21 10:00", "end_time": "29.01.21 22:30","Day": 26},


   {"name": "Владимир Русанов", "imo": 9750701, "ice_class": "Arc7", "speed": 19,
    "start_point": "Саббета 1", "end_point": "точка в Баренцевом море",
    "start_time": "01.01.21 0:00", "end_time": "03.01.21 12:45","Day": 1},


   {"name": "Арктика-2", "imo": 9243801, "ice_class": "Arc5", "speed": 14,
    "start_point": "точка в Баренцевом море", "end_point": "Саббета 2",
    "start_time": "13.01.21 1:30", "end_time": "16.01.21 9:30","Day": 13},


   {"name": "Арктика-2", "imo": 9243801, "ice_class": "Arc5", "speed": 14,
    "start_point": "Саббета 2", "end_point": "точка в Баренцевом море",
    "start_time": "24.01.21 21:00", "end_time": "28.01.21 21:00","Day": 24},


   {"name": "РЗК Константа", "imo": 8711289, "ice_class": "Arc5", "speed": 15,
    "start_point": "Сабетта 2", "end_point": "точка в Баренцевом море",
    "start_time": "29.12.20 15:00", "end_time": "02.01.21 19:30","Day": 29},


   {"name": "РЗК Константа", "imo": 8711289, "ice_class": "Arc5", "speed": 15,
    "start_point": "точка в Баренцевом море", "end_point": "Сабетта 1",
    "start_time": "09.01.21 3:00", "end_time": "13.01.21 21:30","Day": 9},


   {"name": "РЗК Константа", "imo": 8711289, "ice_class": "Arc5", "speed": 15,
    "start_point": "Сабетта 1", "end_point": "точка в Баренцевом море",
    "start_time": "19.01.21 4:00", "end_time": "22.01.21 18:00","Day": 19},


   {"name": "Беринг", "imo": 9267297, "ice_class": "Arc4", "speed": 14,
    "start_point": "Сабетта 2", "end_point": "точка в Баренцевом море",
    "start_time": "05.01.21 21:00", "end_time": "08.01.21 10:00","Day": 5},


   {"name": "Беринг", "imo": 9267297, "ice_class": "Arc4", "speed": 14,
    "start_point": "точка в Баренцевом море", "end_point": "Сабетта 2",
    "start_time": "24.01.21 10:00", "end_time": "28.01.21 22:00","Day": 24},


   {"name": "Северный проект", "imo": 9202053, "ice_class": "Arc4", "speed": 12,
    "start_point": "точка в Баренцевом море", "end_point": "Сабетта 2",
    "start_time": "23.01.21 0:30", "end_time": "30.01.21 21:00","Day": 23},


   {"name": "Мыс Желания", "imo": 9366110, "ice_class": "Arc4", "speed": 16,
    "start_point": "точка в Баренцевом море", "end_point": "Сабетта 2",
    "start_time": "31.12.20 12:00", "end_time": "02.01.21 12:00","Day": 31},


   {"name": "Мыс Желания", "imo": 9366110, "ice_class": "Arc4", "speed": 16,
    "start_point": "Сабетта 2", "end_point": "точка в Баренцевом море",
    "start_time": "16.01.21 0:00", "end_time": "20.01.21 12:00","Day": 16},
   {
       "name": "Никифор Бегичев", "imo": 9014896,"ice_class": "Arc4","speed": 14,"start_point": "точка в Баренцевом море","end_point": "Сабетта 2",
       "start_time": "16.01.21 3:00",
       "end_time": "20.01.21 3:00",
       "Day": 16
   },
   {
       "name": "Никифор Бегичев",
       "imo": 9014896,
       "ice_class": "Arc4",
       "speed": 14,
       "start_point": "Сабетта 2",
       "end_point": "точка в Баренцевом море",
       "start_time": "22.01.21 3:00",
       "end_time": "26.01.21 22:00",
       "Day": 22
   },
   {
       "name": "LNG PHECDA",
       "imo": 9834313,
       "ice_class": "No ice class",
       "speed": 18,
       "start_point": "точка в Баренцевом море",
       "end_point": "Сабетта 1",
       "start_time": "02.01.21 12:30",
       "end_time": "04.01.21 12:00",
       "Day": 2
   },
   {
       "name": "LNG PHECDA",
       "imo": 9834313,
       "ice_class": "No ice class",
       "speed": 18,
       "start_point": "Сабетта 1",
       "end_point": "точка в Баренцевом море",
       "start_time": "04.01.21 11:00",
       "end_time": "06.01.21 14:00",
       "Day": 4
   },
   {
       "name": "LNG PHECDA",
       "imo": 9834313,
       "ice_class": "No ice class",
       "speed": 18,
       "start_point": "точка в Баренцевом море",
       "end_point": "Сабетта 1",
       "start_time": "16.01.21 9:00",
       "end_time": "18.01.21 16:30",
       "Day": 16
   },
   {
       "name": "LNG PHECDA",
       "imo": 9834313,
       "ice_class": "No ice class",
       "speed": 18,
       "start_point": "Сабетта 1",
       "end_point": "точка в Баренцевом море",
       "start_time": "20.01.21 9:00",
       "end_time": "22.01.21 22:00",
       "Day": 20
   },
   {
       "name": "LNG DUBHE",
       "imo": 9834296,
       "ice_class": "No ice class",
       "speed": 16,
       "start_point": "точка в Баренцевом море",
       "end_point": "Сабетта 1",
       "start_time": "02.01.21 0:30",
       "end_time": "03.01.21 23:30",
       "Day": 2
   },
   {
       "name": "LNG DUBHE",
       "imo": 9834296,
       "ice_class": "No ice class",
       "speed": 16,
       "start_point": "Сабетта 1",
       "end_point": "точка в Баренцевом море",
       "start_time": "06.01.21 15:00",
       "end_time": "08.01.21 16:00",
       "Day": 6
   },
   {
       "name": "LNG DUBHE",
       "imo": 9834296,
       "ice_class": "No ice class",
       "speed": 16,
       "start_point": "точка в Баренцевом море",
       "end_point": "Сабетта 1",
       "start_time": "12.01.21 10:00",
       "end_time": "14.01.21 16:30",
       "Day": 12
   },
   {
       "name": "LNG DUBHE",
       "imo": 9834296,
       "ice_class": "No ice class",
       "speed": 16,
       "start_point": "Сабетта 1",
       "end_point": "точка в Баренцевом море",
       "start_time": "13.01.21 19:30",
       "end_time": "15.01.21 10:00",
       "Day": 13
   },
   {
       "name": "LNG DUBHE",
       "imo": 9834296,
       "ice_class": "No ice class",
       "speed": 16,
       "start_point": "точка в Баренцевом море",
       "end_point": "Сабетта 1",
       "start_time": "22.01.21 7:00",
       "end_time": "24.01.21 18:30",
       "Day": 22
   },
   {
       "name": "LNG DUBHE",
       "imo": 9834296,
       "ice_class": "No ice class",
       "speed": 16,
       "start_point": "Сабетта 1",
       "end_point": "точка в Баренцевом море",
       "start_time": "23.01.21 0:30",
       "end_time": "25.01.21 23:30",
       "Day": 23
   }
]


wb = Workbook()
ws = wb.active
columns = ["Судно", "Номер Заявки", "Ледокол", "Дата и время начала", "Дата и время окончания", "Отрезок начала", "Отрезок окончания"]
ws.append(columns)


def generate_start(start_point):
   generated_point = random.choice([i for i in range(1, 11)])
   edges = list(range(10, generated_point - 1, -1))
   if "Саббета 1" in [start_point]:
       return 14
   elif "Саббета 2" in [start_point]:
       return 13
   elif "Саббета 3" in [start_point]:
       return 11
   else:
       return generated_point


def generate_end(end_point):
   generated_point = random.choice([i for i in range(1, 11)])
   edges = list(range(10, generated_point - 1, -1))
   if "Саббета 1" in [end_point]:
       return 14
   elif "Саббета 2" in [end_point]:
       return 13
   elif "Саббета 3" in [end_point]:
       return 11
   else:
       return generated_point




def SamSip (ice_class,start_point):
   if ice_class in ["Arc7","Arc8","Arc9"]:
       return " - "
   elif start_point == "Саббетта 3":
       return icebreakers[random.choice([i for i in range(3, 4)])]
   else:
       return icebreakers[random.choice([i for i in range(0,4)])]


# ЗАПОЛНЕНИЕ
request_counter = 1
for i, request in enumerate(requests):
   ship_name = request["name"]
   request_number = request_counter
   start_time = request["start_time"]
   end_time = request["end_time"]


   request_counter += 1  # Увеличиваем счетчик заявок
   # Заполняем столбцы начальный и конечный отрезок
   row = [ship_name, request_number, SamSip(request["ice_class"],request["start_point"]), start_time, end_time, generate_start(request["start_point"]),generate_end(request["end_point"])]
   ws.append(row)


#СТИЛЬ
for column_cells in ws.columns:
   max_length = 0
   column = column_cells[0].column_letter
   for cell in column_cells:
       try:
           if len(str(cell.value)) > max_length:
               max_length = len(cell.value)
       except:
           pass
   adjusted_width = (max_length + 3)
   ws.column_dimensions[column].width = adjusted_width
border = Border(
   left=Side(style='thin'),
   right=Side(style='thin'),
   top=Side(style='thin'),
   bottom=Side(style='thin')
)


# Проходим по всем строкам и столбцам, применяя стиль границы к каждой ячейке
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
   for cell in row:
       cell.border = border


#СОХРАНЕНИЕ
# Сохраняем результаты в Excel
wb.save("optimal_routes.xlsx")


# Открываем файл с помощью системного приложения по умолчанию
subprocess.Popen(["start", "optimal_routes.xlsx"], shell=True)



